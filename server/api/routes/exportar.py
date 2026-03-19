"""
server/api/routes/exportar.py
==============================
Exportación de reportes a Excel (.xlsx) y PDF.

Endpoints:
  GET /api/exportar/excel/mensual   → Reporte mensual en Excel
  GET /api/exportar/excel/alumno    → Historial de un alumno en Excel
  GET /api/exportar/pdf/diario      → Reporte del día en PDF simple

Dependencias opcionales (se cargan solo si están instaladas):
  pip install openpyxl reportlab

Si no están instaladas, los endpoints devuelven 501 con instrucciones.
"""

import io
import logging
from datetime import datetime, date, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, func
from sqlalchemy.orm import Session

from server.db.database import get_db
from server.db import models
from server.db.models import RolUsuario, TipoEvento, EstadoAsistencia
from server.core.security import require_rol

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/exportar", tags=["Exportación"])


def _verificar_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="openpyxl no instalado. Ejecutar: pip install openpyxl"
        )


def _verificar_reportlab():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        return True
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="reportlab no instalado. Ejecutar: pip install reportlab"
        )


# ================================================================== #
# EXCEL: Reporte mensual completo
# ================================================================== #

@router.get("/excel/mensual", summary="Reporte mensual en Excel")
def exportar_excel_mensual(
    año: int  = Query(default=datetime.utcnow().year),
    mes: int  = Query(default=datetime.utcnow().month, ge=1, le=12),
    db:  Session = Depends(get_db),
    _user = Depends(require_rol(RolUsuario.ADMIN, RolUsuario.TUTOR)),
):
    """
    Genera un archivo Excel con la asistencia de todos los alumnos del mes.
    Cada fila = un alumno. Cada columna = un día del mes.
    Celda verde = PRESENTE, rojo = AUSENTE, amarillo = TARDANZA, gris = FIN DE SEMANA.
    """
    openpyxl = _verificar_openpyxl()
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # --- Rango del mes ---
    inicio = datetime(año, mes, 1)
    if mes == 12:
        fin = datetime(año + 1, 1, 1)
    else:
        fin = datetime(año, mes + 1, 1)

    # --- Obtener días del mes ---
    dias = []
    d = inicio
    while d < fin:
        dias.append(d)
        d += timedelta(days=1)

    # --- Obtener alumnos activos ---
    alumnos = db.query(models.Alumno).filter(
        models.Alumno.activo == True
    ).order_by(models.Alumno.grado, models.Alumno.seccion, models.Alumno.apellidos).all()

    # --- Obtener todos los registros del mes (una sola query) ---
    registros = db.query(models.Asistencia).filter(
        and_(
            models.Asistencia.fecha >= inicio,
            models.Asistencia.fecha < fin,
            models.Asistencia.tipo_evento == TipoEvento.ENTRADA,
        )
    ).all()

    # Mapa: (alumno_id, fecha_str) → estado
    mapa = {}
    for r in registros:
        key = (r.alumno_id, r.fecha.strftime("%Y-%m-%d"))
        mapa[key] = r.estado.value if hasattr(r.estado, 'value') else r.estado

    # --- Crear workbook ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Asistencia {año}-{mes:02d}"

    # Colores
    VERDE    = PatternFill("solid", fgColor="C6EFCE")
    ROJO     = PatternFill("solid", fgColor="FFC7CE")
    AMARILLO = PatternFill("solid", fgColor="FFEB9C")
    GRIS     = PatternFill("solid", fgColor="D9D9D9")
    AZUL     = PatternFill("solid", fgColor="BDD7EE")

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- Fila de título ---
    ws.merge_cells(f"A1:{get_column_letter(4 + len(dias))}1")
    titulo = ws["A1"]
    titulo.value = f"REPORTE DE ASISTENCIA — {inicio.strftime('%B %Y').upper()}"
    titulo.font  = Font(bold=True, size=14)
    titulo.alignment = center
    titulo.fill  = AZUL
    ws.row_dimensions[1].height = 30

    # --- Fila de encabezados ---
    ws.row_dimensions[2].height = 45
    headers = ["N°", "Código", "Apellidos y Nombres", "Grado"] + \
              [d.strftime("%d\n%a") for d in dias] + ["TOTAL\nPRESENTE", "% ASIST."]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = bold
        c.alignment = center
        c.fill = AZUL
        c.border = thin

    # Anchos de columna
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 8
    for i, d in enumerate(dias, 5):
        ws.column_dimensions[get_column_letter(i)].width = 5

    # Calcular días lectivos (L-V)
    dias_lectivos = sum(1 for d in dias if d.weekday() < 5)

    # --- Filas de alumnos ---
    for fila_n, alumno in enumerate(alumnos, 1):
        row = fila_n + 2
        ws.row_dimensions[row].height = 18

        ws.cell(row=row, column=1, value=fila_n).border = thin
        ws.cell(row=row, column=2, value=alumno.codigo).border = thin
        c = ws.cell(row=row, column=3, value=alumno.nombre_completo())
        c.border = thin
        ws.cell(row=row, column=4, value=f"{alumno.grado}{alumno.seccion}").border = thin

        dias_presentes = 0
        for col_n, dia in enumerate(dias, 5):
            c = ws.cell(row=row, column=col_n)
            c.alignment = center
            c.border = thin

            if dia.weekday() >= 5:  # Sábado/Domingo
                c.fill = GRIS
                c.value = "—"
                continue

            estado = mapa.get((alumno.id, dia.strftime("%Y-%m-%d")))
            if estado == "PRESENTE":
                c.fill = VERDE
                c.value = "P"
                dias_presentes += 1
            elif estado == "TARDANZA":
                c.fill = AMARILLO
                c.value = "T"
                dias_presentes += 1  # Tardanza cuenta como presente
            elif estado == "JUSTIFICADO":
                c.fill = AZUL
                c.value = "J"
                dias_presentes += 1
            elif estado == "AUSENTE":
                c.fill = ROJO
                c.value = "A"
            else:
                c.fill = ROJO
                c.value = "A"  # Sin registro = ausente

        # Totales
        col_total = 5 + len(dias)
        ws.cell(row=row, column=col_total, value=dias_presentes).border = thin
        pct = round(dias_presentes / dias_lectivos * 100, 1) if dias_lectivos > 0 else 0
        c_pct = ws.cell(row=row, column=col_total + 1, value=f"{pct}%")
        c_pct.border = thin
        if pct < 85:
            c_pct.fill = ROJO
        elif pct < 95:
            c_pct.fill = AMARILLO

    # --- Leyenda ---
    leyenda_row = len(alumnos) + 4
    ws.cell(row=leyenda_row, column=1, value="LEYENDA:").font = bold
    leyendas = [("P", "Presente", VERDE), ("T", "Tardanza", AMARILLO),
                ("A", "Ausente", ROJO), ("J", "Justificado", AZUL)]
    for i, (letra, desc, fill) in enumerate(leyendas):
        c1 = ws.cell(row=leyenda_row, column=2 + i * 2, value=letra)
        c1.fill = fill
        c1.alignment = center
        ws.cell(row=leyenda_row, column=3 + i * 2, value=desc)

    # --- Serializar a bytes ---
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre_archivo = f"asistencia_{año}_{mes:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )


# ================================================================== #
# EXCEL: Historial de un alumno
# ================================================================== #

@router.get("/excel/alumno/{alumno_id}", summary="Historial de asistencia de un alumno en Excel")
def exportar_excel_alumno(
    alumno_id: int,
    dias:      int = Query(default=30, le=365),
    db:        Session = Depends(get_db),
    _user      = Depends(require_rol(RolUsuario.ADMIN, RolUsuario.TUTOR)),
):
    openpyxl = _verificar_openpyxl()
    from openpyxl.styles import PatternFill, Font, Alignment

    alumno = db.get(models.Alumno, alumno_id)
    if not alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")

    desde = datetime.utcnow() - timedelta(days=dias)
    registros = db.query(models.Asistencia).filter(
        and_(
            models.Asistencia.alumno_id == alumno_id,
            models.Asistencia.fecha >= desde,
        )
    ).order_by(models.Asistencia.fecha.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"

    bold = Font(bold=True)
    # Encabezado del alumno
    ws["A1"] = f"Alumno: {alumno.nombre_completo()}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Grado: {alumno.grado}{alumno.seccion} — Turno: {alumno.turno}"
    ws["A3"] = f"Período: últimos {dias} días"

    # Cabecera de tabla
    headers = ["Fecha", "Día", "Tipo Evento", "Estado", "Hora", "Confianza IA", "Registrado por"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = bold

    COLORES = {
        "ENTRADA": "C6EFCE",
        "SALIDA":  "BDD7EE",
        "TARDANZA": "FFEB9C",
        "AUSENTE": "FFC7CE",
    }

    for i, r in enumerate(registros, 6):
        tipo = r.tipo_evento.value if hasattr(r.tipo_evento, 'value') else r.tipo_evento
        estado = r.estado.value if hasattr(r.estado, 'value') else r.estado
        fill = PatternFill("solid", fgColor=COLORES.get(tipo, "FFFFFF"))

        row_data = [
            r.fecha.strftime("%Y-%m-%d"),
            r.fecha.strftime("%A"),
            tipo,
            estado,
            r.fecha.strftime("%H:%M:%S"),
            f"{r.confianza:.0%}" if r.confianza else "Manual",
            r.registrado_por,
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = fill

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = f"historial_{alumno.codigo}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )


# ================================================================== #
# PDF: Reporte del día
# ================================================================== #

@router.get("/pdf/diario", summary="Reporte del día en PDF")
def exportar_pdf_diario(
    fecha: Optional[date] = Query(None),
    db:    Session = Depends(get_db),
    _user  = Depends(require_rol(RolUsuario.ADMIN, RolUsuario.TUTOR)),
):
    """
    Genera un PDF con la lista de presentes y ausentes del día.
    Listo para imprimir o enviar por correo.
    """
    _verificar_reportlab()
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    fecha_consulta = fecha or date.today()
    inicio = datetime.combine(fecha_consulta, datetime.min.time())
    fin    = datetime.combine(fecha_consulta, datetime.max.time())

    # Registros del día
    registros = db.query(models.Asistencia).filter(
        and_(
            models.Asistencia.fecha.between(inicio, fin),
            models.Asistencia.tipo_evento == TipoEvento.ENTRADA,
        )
    ).order_by(models.Asistencia.fecha).all()

    presentes_ids = {r.alumno_id for r in registros}

    # Ausentes
    ausentes = db.query(models.Alumno).filter(
        models.Alumno.activo == True,
        ~models.Alumno.id.in_(presentes_ids),
    ).order_by(models.Alumno.grado, models.Alumno.apellidos).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    story = []

    # Título
    story.append(Paragraph(
        f"<b>REPORTE DE ASISTENCIA DIARIA</b>", styles["Title"]
    ))
    story.append(Paragraph(
        f"Fecha: {fecha_consulta.strftime('%d de %B de %Y')} | "
        f"Presentes: {len(presentes_ids)} | Ausentes: {len(ausentes)}",
        styles["Normal"]
    ))
    story.append(Spacer(1, 12))

    # Tabla de presentes
    story.append(Paragraph("<b>PRESENTES</b>", styles["Heading2"]))
    if registros:
        data = [["N°", "Código", "Alumno", "Grado", "Hora", "Estado"]]
        for i, r in enumerate(registros, 1):
            a = r.alumno
            estado = r.estado.value if hasattr(r.estado, 'value') else str(r.estado)
            data.append([
                str(i),
                a.codigo if a else "—",
                a.nombre_completo() if a else "—",
                f"{a.grado}{a.seccion}" if a else "—",
                r.fecha.strftime("%H:%M"),
                estado,
            ])

        t = Table(data, colWidths=[25, 60, 180, 45, 45, 70])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a5276")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF5FB")]),
            ("BACKGROUND", (5, 1), (5, -1), colors.HexColor("#D5F5E3")),
            ("GRID",       (0, 0), (-1, -1), 0.3, colors.grey),
            ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(t)

    story.append(Spacer(1, 20))

    # Tabla de ausentes
    story.append(Paragraph("<b>AUSENTES</b>", styles["Heading2"]))
    if ausentes:
        data2 = [["N°", "Código", "Alumno", "Grado", "Turno"]]
        for i, a in enumerate(ausentes, 1):
            data2.append([str(i), a.codigo, a.nombre_completo(), f"{a.grado}{a.seccion}", a.turno])

        t2 = Table(data2, colWidths=[25, 60, 220, 45, 60])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#922b21")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FDEDEC")]),
            ("GRID",       (0, 0), (-1, -1), 0.3, colors.grey),
            ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(t2)

    # Pie de página
    story.append(Spacer(1, 30))
    story.append(Paragraph(
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Sistema Colegio Asistencia v1.0",
        styles["Normal"]
    ))

    doc.build(story)
    buf.seek(0)

    nombre = f"reporte_diario_{fecha_consulta.strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )


# ================================================================== #
# PDF: Reporte individual por alumno
# ================================================================== #

@router.get("/pdf/alumno/{alumno_id}", summary="Reporte individual de alumno en PDF")
def exportar_pdf_alumno(
    alumno_id: int,
    dias:      int = Query(default=30, le=180, description="Días de historial"),
    db:        Session = Depends(get_db),
    _user      = Depends(require_rol(RolUsuario.ADMIN, RolUsuario.TUTOR)),
):
    """
    Genera un PDF con el reporte individual completo de un alumno.
    Incluye: datos personales, estadísticas del período y tabla de registros.
    Ideal para reuniones con padres o para adjuntar a una justificación.
    """
    _verificar_reportlab()
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    alumno = db.get(models.Alumno, alumno_id)
    if not alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")

    desde = datetime.utcnow() - timedelta(days=dias)

    registros = db.query(models.Asistencia).filter(
        and_(
            models.Asistencia.alumno_id == alumno_id,
            models.Asistencia.fecha >= desde,
            models.Asistencia.tipo_evento == TipoEvento.ENTRADA,
        )
    ).order_by(models.Asistencia.fecha.desc()).all()

    # Calcular estadísticas
    total_dias = dias
    dias_presente = len(registros)
    tardanzas     = sum(1 for r in registros if r.estado == EstadoAsistencia.TARDANZA)
    justificados  = sum(1 for r in registros if r.estado == EstadoAsistencia.JUSTIFICADO)
    ausencias     = max(0, total_dias - dias_presente)
    pct           = round(dias_presente / total_dias * 100, 1) if total_dias > 0 else 0

    # Config PDF
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
        leftMargin=2*cm, rightMargin=2*cm,
    )
    styles = getSampleStyleSheet()
    story  = []

    # Cabecera institucional
    cfg_nombre = db.query(models.Configuracion).filter(
        models.Configuracion.clave == "nombre_colegio"
    ).first()
    nombre_colegio = cfg_nombre.valor if cfg_nombre else "Colegio"

    story.append(Paragraph(
        f"<b>{nombre_colegio.upper()}</b>", styles["Title"]
    ))
    story.append(Paragraph(
        "REPORTE INDIVIDUAL DE ASISTENCIA", styles["Heading2"]
    ))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1a5276")))
    story.append(Spacer(1, 0.4*cm))

    # Datos del alumno (tabla 2 columnas)
    datos_alumno = [
        ["Apellidos y Nombres:", alumno.nombre_completo()],
        ["Código:",              alumno.codigo],
        ["Grado / Sección:",    f"{alumno.grado}° {alumno.seccion}"],
        ["Turno:",               alumno.turno],
        ["Período analizado:",  f"Últimos {dias} días ({desde.strftime('%d/%m/%Y')} — {datetime.utcnow().strftime('%d/%m/%Y')})"],
    ]
    t_datos = Table(datos_alumno, colWidths=[5*cm, 12*cm])
    t_datos.setStyle(TableStyle([
        ("FONTNAME",  (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE",  (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1a5276")),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t_datos)
    story.append(Spacer(1, 0.6*cm))

    # Estadísticas en cajas de color
    stats_data = [
        ["DÍAS\nPRESENTE", "AUSENCIAS", "TARDANZAS", "JUSTIFICADOS", "% ASISTENCIA"],
        [str(dias_presente), str(ausencias), str(tardanzas), str(justificados), f"{pct}%"],
    ]
    t_stats = Table(stats_data, colWidths=[3.3*cm]*5)
    color_pct = colors.HexColor("#155724") if pct >= 90 else (
        colors.HexColor("#856404") if pct >= 75 else colors.HexColor("#721c24")
    )
    t_stats.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1a5276")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 8),
        ("FONTNAME",    (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 1), (-1, 1), 20),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND",  (4, 1), (4, 1), color_pct),
        ("TEXTCOLOR",   (4, 1), (4, 1), colors.white),
    ]))
    story.append(t_stats)
    story.append(Spacer(1, 0.6*cm))

    # Tabla de registros
    story.append(Paragraph("<b>DETALLE DE REGISTROS</b>", styles["Heading3"]))
    story.append(Spacer(1, 0.2*cm))

    if registros:
        ESTADO_COLOR = {
            "PRESENTE":    colors.HexColor("#d4edda"),
            "TARDANZA":    colors.HexColor("#fff3cd"),
            "JUSTIFICADO": colors.HexColor("#cce5ff"),
            "AUSENTE":     colors.HexColor("#f8d7da"),
        }
        data_reg = [["N°", "Fecha", "Día", "Hora", "Estado", "Registrado por"]]
        for i, r in enumerate(registros[:60], 1):  # Máximo 60 registros en el PDF
            estado_str = r.estado.value if hasattr(r.estado, 'value') else str(r.estado)
            data_reg.append([
                str(i),
                r.fecha.strftime("%d/%m/%Y"),
                r.fecha.strftime("%A")[:3],
                r.fecha.strftime("%H:%M"),
                estado_str,
                r.registrado_por[:20] if r.registrado_por else "—",
            ])

        t_reg = Table(data_reg, colWidths=[1*cm, 2.5*cm, 1.5*cm, 1.5*cm, 2.5*cm, 7.5*cm])
        style_cmds = [
            ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1a5276")),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 8),
            ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
            ("GRID",        (0, 0), (-1, -1), 0.3, colors.grey),
            ("TOPPADDING",  (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        # Colorear filas según estado
        for row_i, r in enumerate(registros[:60], 1):
            estado_str = r.estado.value if hasattr(r.estado, 'value') else str(r.estado)
            bg = ESTADO_COLOR.get(estado_str, colors.white)
            style_cmds.append(("BACKGROUND", (0, row_i), (-1, row_i), bg))

        t_reg.setStyle(TableStyle(style_cmds))
        story.append(t_reg)

        if len(registros) > 60:
            story.append(Paragraph(
                f"<i>Se muestran los 60 registros más recientes de {len(registros)} totales.</i>",
                styles["Normal"]
            ))
    else:
        story.append(Paragraph("Sin registros en el período seleccionado.", styles["Normal"]))

    # Pie de página
    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    story.append(Paragraph(
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | "
        f"Sistema Colegio Asistencia v1.0 | "
        f"Documento confidencial — uso interno",
        ParagraphStyle("pie", parent=styles["Normal"], fontSize=7, textColor=colors.grey)
    ))

    doc.build(story)
    buf.seek(0)

    nombre = f"reporte_{alumno.codigo}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={nombre}"},
    )
