"""
tests/server/test_exportar.py
==============================
Tests de los endpoints de exportación Excel y PDF.

Verifica que los endpoints devuelvan los content-types correctos,
que funcionen con datos vacíos y que manejen correctamente
los casos donde openpyxl/reportlab no están instalados.

Uso:
    pytest tests/server/test_exportar.py -v
"""

import pytest
import io
from datetime import datetime
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from fastapi.testclient import TestClient
from unittest.mock import patch, MagicMock

from server.db.database import Base, get_db
from server.db import models
from server.core.security import hash_password
from server.main import app

# ------------------------------------------------------------------ #
# Setup DB en memoria
# ------------------------------------------------------------------ #
engine_test = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
TestingSession = sessionmaker(bind=engine_test)


def override_get_db():
    db = TestingSession()
    try:
        yield db
    finally:
        db.close()


app.dependency_overrides[get_db] = override_get_db


@pytest.fixture(autouse=True)
def setup_db():
    Base.metadata.create_all(bind=engine_test)
    db = TestingSession()
    try:
        # Admin
        db.add(models.UsuarioSistema(
            username="admin_t", password_hash=hash_password("admin1234"),
            nombre_display="Admin", rol=models.RolUsuario.ADMIN,
        ))
        # Alumno con registros
        alumno = models.Alumno(
            codigo="EXP001", nombres="Carlos", apellidos="Mendoza Test",
            grado="2", seccion="B", turno="MAÑANA", activo=True,
        )
        db.add(alumno)
        db.flush()

        # Registros de asistencia del mes actual
        ahora = datetime.utcnow()
        for dia in [1, 2, 3, 5, 6]:  # Algunos días del mes
            fecha = ahora.replace(day=dia, hour=8, minute=5, second=0)
            db.add(models.Asistencia(
                alumno_id=alumno.id,
                fecha=fecha,
                tipo_evento=models.TipoEvento.ENTRADA,
                estado=models.EstadoAsistencia.PRESENTE,
                registrado_por="test",
            ))
        # Un día de tardanza
        fecha_tard = ahora.replace(day=7, hour=8, minute=30, second=0)
        db.add(models.Asistencia(
            alumno_id=alumno.id, fecha=fecha_tard,
            tipo_evento=models.TipoEvento.ENTRADA,
            estado=models.EstadoAsistencia.TARDANZA,
            registrado_por="test",
        ))
        db.commit()
    finally:
        db.close()
    yield
    Base.metadata.drop_all(bind=engine_test)


@pytest.fixture
def client():
    return TestClient(app)


@pytest.fixture
def token_admin(client):
    r = client.post("/api/auth/login", json={"username": "admin_t", "password": "admin1234"})
    return r.json()["access_token"]


@pytest.fixture
def headers(token_admin):
    return {"Authorization": f"Bearer {token_admin}"}


# ================================================================== #
# TEST: Excel mensual
# ================================================================== #

class TestExcelMensual:

    def test_devuelve_content_type_excel(self, client, headers):
        """El endpoint debe devolver un archivo .xlsx válido."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl no instalado")

        ahora = datetime.utcnow()
        r = client.get(
            f"/api/exportar/excel/mensual?año={ahora.year}&mes={ahora.month}",
            headers=headers,
        )
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]
        assert r.headers["content-disposition"].startswith("attachment")
        assert len(r.content) > 1000  # El Excel no debe estar vacío

    def test_excel_se_puede_abrir_con_openpyxl(self, client, headers):
        """El archivo generado debe ser un Excel válido que openpyxl puede leer."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl no instalado")

        ahora = datetime.utcnow()
        r = client.get(
            f"/api/exportar/excel/mensual?año={ahora.year}&mes={ahora.month}",
            headers=headers,
        )
        assert r.status_code == 200

        # Intentar abrir el Excel generado
        buf = io.BytesIO(r.content)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        # Debe tener al menos la fila de título y la de headers
        assert ws.max_row >= 2
        # El título debe mencionar el mes
        titulo_celda = ws["A1"].value
        assert titulo_celda and "ASISTENCIA" in titulo_celda.upper()

    def test_sin_openpyxl_retorna_501(self, client, headers):
        """Sin openpyxl instalado, el endpoint debe devolver 501."""
        with patch("server.api.routes.exportar._verificar_openpyxl") as mock:
            from fastapi import HTTPException
            mock.side_effect = HTTPException(status_code=501, detail="openpyxl no instalado")
            r = client.get("/api/exportar/excel/mensual?año=2025&mes=1", headers=headers)
            assert r.status_code == 501

    def test_requiere_autenticacion(self, client):
        r = client.get("/api/exportar/excel/mensual?año=2025&mes=1")
        assert r.status_code == 401


# ================================================================== #
# TEST: Excel por alumno
# ================================================================== #

class TestExcelAlumno:

    def test_excel_alumno_valido(self, client, headers):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl no instalado")

        # ID 1 = primer alumno creado en setup
        r = client.get("/api/exportar/excel/alumno/1?dias=30", headers=headers)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]

    def test_excel_alumno_inexistente(self, client, headers):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl no instalado")

        r = client.get("/api/exportar/excel/alumno/99999?dias=30", headers=headers)
        assert r.status_code == 404

    def test_excel_alumno_sin_registros(self, client, headers):
        """Un alumno sin registros no debe causar error, solo un Excel vacío."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl no instalado")

        # Crear alumno sin registros
        db = TestingSession()
        nuevo = models.Alumno(
            codigo="VACIO01", nombres="Sin", apellidos="Registros",
            grado="1", seccion="A", turno="MAÑANA",
        )
        db.add(nuevo)
        db.commit()
        nuevo_id = nuevo.id
        db.close()

        r = client.get(f"/api/exportar/excel/alumno/{nuevo_id}?dias=30", headers=headers)
        assert r.status_code == 200


# ================================================================== #
# TEST: PDF diario
# ================================================================== #

class TestPDFDiario:

    def test_devuelve_content_type_pdf(self, client, headers):
        """El endpoint debe devolver un PDF válido."""
        try:
            from reportlab.lib.pagesizes import A4
        except ImportError:
            pytest.skip("reportlab no instalado")

        from datetime import date
        hoy = date.today().isoformat()
        r = client.get(f"/api/exportar/pdf/diario?fecha={hoy}", headers=headers)
        assert r.status_code == 200
        assert r.headers["content-type"] == "application/pdf"
        # PDF empieza con "%PDF"
        assert r.content[:4] == b"%PDF"

    def test_pdf_con_fecha_pasada(self, client, headers):
        """El PDF para una fecha pasada sin datos debe generarse igual."""
        try:
            from reportlab.lib.pagesizes import A4
        except ImportError:
            pytest.skip("reportlab no instalado")

        r = client.get("/api/exportar/pdf/diario?fecha=2020-01-15", headers=headers)
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"

    def test_sin_reportlab_retorna_501(self, client, headers):
        """Sin reportlab instalado, el endpoint debe devolver 501."""
        with patch("server.api.routes.exportar._verificar_reportlab") as mock:
            from fastapi import HTTPException
            mock.side_effect = HTTPException(status_code=501, detail="reportlab no instalado")
            r = client.get("/api/exportar/pdf/diario", headers=headers)
            assert r.status_code == 501

    def test_pdf_requiere_autenticacion(self, client):
        r = client.get("/api/exportar/pdf/diario")
        assert r.status_code == 401


# ================================================================== #
# TEST: Permisos de exportación
# ================================================================== #

class TestPermisosExportacion:

    @pytest.fixture
    def token_portero(self, client):
        """Crear y autenticar un portero."""
        db = TestingSession()
        db.add(models.UsuarioSistema(
            username="portero_exp", password_hash=hash_password("p1234567"),
            nombre_display="Portero", rol=models.RolUsuario.PORTERO,
        ))
        db.commit()
        db.close()
        r = client.post("/api/auth/login", json={"username": "portero_exp", "password": "p1234567"})
        return r.json()["access_token"]

    def test_portero_no_puede_exportar_excel(self, client, token_portero):
        """Los porteros no tienen acceso a exportación."""
        headers = {"Authorization": f"Bearer {token_portero}"}
        r = client.get("/api/exportar/excel/mensual?año=2025&mes=1", headers=headers)
        assert r.status_code == 403

    def test_admin_puede_exportar_todo(self, client, headers):
        """El admin puede acceder a todos los endpoints de exportación."""
        # Solo verificamos que no retorne 403 (puede retornar 501 si falta librería)
        endpoints = [
            "/api/exportar/excel/mensual?año=2025&mes=1",
            "/api/exportar/excel/alumno/1?dias=7",
            "/api/exportar/pdf/diario?fecha=2025-01-15",
        ]
        for ep in endpoints:
            r = client.get(ep, headers=headers)
            assert r.status_code not in [401, 403], f"Acceso denegado en {ep}"
